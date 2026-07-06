"""흐름 트리(Flow Tree) → QA 양식 Excel 그리드 렌더러.

PoC(docs/PoC_FlowTree_doc82_v3.xlsx)의 trie→grid 렌더링을 정식화.
노드를 좌→우 트라이로 펼쳐, 각 노드를 (타입, 내용) 열 쌍으로 배치한다.
설계: docs/flow_tree_schema_design.md §6.4
"""
import io
import re
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 타입별 셀 배경색
TYPE_FILL = {
    "PR": "fef3c7", "C": "dbeafe", "D": "dcfce7", "T": "fae8ff",
    "H": "e0e7ff", "V": "fee2e2", "RC": "dbeafe", "DC": "dbeafe", "DD": "dbeafe",
}
# 범례 (사람 골든 기준)
LEGEND = [
    ("D", "Display"), ("C", "Click"), ("RC", "우클릭"), ("DC", "더블클릭"),
    ("T", "input"), ("H", "Hover"), ("DD", "Drag Down"),
    ("PR", "Pre-condition"), ("V", "Verify"),
]
# D(표시) 뒤에 와도 같은 행에 못 이어붙는 "진짜 새 액션" (R2-1). V는 결과(D처럼)로 취급해 제외.
_ROW_BREAKING_ACTIONS = {"C", "DC", "H", "T", "RC", "DD"}


_FLOW_ACTION_TYPES = {"C", "T", "H", "DC", "RC", "DD"}


def _build_tc_from_path(path: List[Dict], tc_id: str) -> Dict:
    """루트→리프 노드 경로에서 TC 한 건을 추출한다."""
    preconditions = [n["content"] for n in path if n.get("type") == "PR" and n.get("content")]
    seq = [n for n in path if n.get("type") != "PR"]

    steps: List[Dict] = []
    entry_displays: List[str] = []
    cur = None
    for n in seq:
        t = n.get("type")
        content = n.get("content", "")
        if t in _FLOW_ACTION_TYPES:
            if cur is None and entry_displays:
                steps.append({"step": len(steps) + 1,
                               "action": "화면/팝업에 진입한다",
                               "expected": "\n".join(entry_displays)})
            elif cur is not None:
                steps.append(cur)
            cur = {"step": len(steps) + 1, "action": content, "expected": ""}
        elif t in ("D", "V") and content:
            if cur is None:
                entry_displays.append(content)
            else:
                cur["expected"] = (cur["expected"] + "\n" + content).strip() if cur["expected"] else content
    if cur is not None:
        steps.append(cur)
    elif entry_displays:
        steps.append({"step": 1, "action": "표시 항목을 확인한다", "expected": "\n".join(entry_displays)})

    expected_result = steps[-1]["expected"] if steps and steps[-1].get("expected") else ""
    if not expected_result:
        all_d = [n.get("content", "") for n in seq if n.get("type") in ("D", "V") and n.get("content")]
        expected_result = all_d[-1] if all_d else ""

    last_action = next((n.get("content", "") for n in reversed(path) if n.get("type") in _FLOW_ACTION_TYPES), "")
    last_d = next((n.get("content", "") for n in reversed(path) if n.get("type") in ("D", "V")), "")
    core = (last_action or last_d or "")[:80]
    title = core + ("…" if len(core) == 80 else "")

    joined = " ".join(n.get("content", "") for n in path)
    tc_type = "negative" if any(k in joined for k in ["오류", "실패", "잘못", "유효하지", "불가", "에러"]) else "positive"

    menu_path = ""
    for n in path:
        if n.get("menu_path"):
            menu_path = n["menu_path"]
            break
    if not menu_path and path:
        menu_path = path[0].get("content", "")

    spec_page = ""
    for n in reversed(path):
        if n.get("spec_page"):
            spec_page = n["spec_page"]
            break

    steps_text = "\n".join(f"{s['step']}. {s['action']}" + (f"\n   → {s['expected']}" if s.get('expected') else "")
                           for s in steps)
    return {
        "tc_id": tc_id,
        "category": menu_path,
        "title": title,
        "tc_type": tc_type,
        "spec_page": spec_page,
        "preconditions": " / ".join(preconditions),
        "steps": steps_text,
        "expected_result": expected_result,
    }


def render_flow_tree_excel(flow_tree: Dict, include_tc: bool = False,
                           tc_data: Optional[Dict[int, Dict]] = None,
                           _out_paths: Optional[dict] = None):
    """흐름 트리 dict → (xlsx BytesIO, path_tracker).

    flow_tree: {"title": str, "tree": [node, ...]}
    include_tc: True이면 TC 컬럼을 오른쪽에 1:1로 붙임.
    tc_data: {row_number: tc_dict} — Gemini 생성 자연어 TC. 없으면 기계적 선형화 사용.
    _out_paths: 호출측이 넘겨준 dict에 path_tracker를 채워 반환 (2단계 Gemini 생성용).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "흐름트리"

    thin = Side(style="thin", color="d1d5db")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # r1: 제목
    title = (flow_tree or {}).get("title", "")
    ws.cell(1, 2, f"흐름 메뉴트리 — {title}").font = Font(bold=True, size=13)

    # r2: 범례
    col = 2
    for code, label in LEGEND:
        c = ws.cell(2, col, code)
        c.font = Font(bold=True, size=9)
        c.fill = PatternFill("solid", fgColor=TYPE_FILL.get(code, "eeeeee"))
        ws.cell(2, col + 1, label).font = Font(size=9, italic=True)
        col += 2

    tree = (flow_tree or {}).get("tree", []) or []
    state = {"row": 4, "max_col": 2}
    # path_tracker: 각 행 번호 → 루트에서 그 행의 마지막 노드까지의 경로
    path_tracker: Dict[int, List[Dict]] = {}

    def write_cell(r: int, depth: int, node: Dict):
        tcol = 2 + 2 * depth
        ccol = tcol + 1
        state["max_col"] = max(state["max_col"], tcol)
        ntype = node.get("type", "?")
        content = node.get("content", "")
        a = ws.cell(r, tcol, ntype)
        a.font = Font(bold=True, size=9, color="374151")
        a.fill = PatternFill("solid", fgColor=TYPE_FILL.get(ntype, "f3f4f6"))
        a.alignment = Alignment(horizontal="center", vertical="top")
        a.border = border
        b = ws.cell(r, ccol, content)
        b.font = Font(size=9)
        b.alignment = Alignment(vertical="top", wrap_text=True)
        b.border = border

    def render(node: Dict, depth: int, path: List[Dict] = None):
        if path is None:
            path = []
        current_path = path + [node]
        write_cell(state["row"], depth, node)
        # 같은 행에 이어지는 노드는 누적(append) — 덮어쓰면 중간 D들이 사라짐
        row = state["row"]
        if row not in path_tracker:
            path_tracker[row] = list(current_path)  # 첫 노드: 조상 맥락 포함 초기화
        else:
            path_tracker[row].append(node)           # 이후 노드: 현재 노드만 추가

        parent_type = node.get("type")
        children = node.get("children") or []
        prev_was_d = False
        next_depth = depth + 1
        for i, ch in enumerate(children):
            ch_type = ch.get("type")
            if i == 0:
                if ch_type == "PR":
                    new_row = True
                elif parent_type == "D" and ch_type in _ROW_BREAKING_ACTIONS:
                    new_row = True
                else:
                    new_row = False
            else:
                new_row = not (prev_was_d and ch_type == "D")
            if new_row:
                state["row"] += 1
                next_depth = depth + 1
            render(ch, next_depth, current_path)
            next_depth += 1
            prev_was_d = (ch_type == "D")

    for root in tree:
        render(root, 0, [])
        state["row"] += 1

    # r3: Step 헤더
    n_steps = (state["max_col"] - 2) // 2
    for k in range(n_steps + 1):
        label = "사전조건" if k == 0 else f"Step{k}"
        c = ws.cell(3, 2 + 2 * k, label)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="374151")
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(2 + 2 * k)].width = 5
        ws.column_dimensions[get_column_letter(3 + 2 * k)].width = 34

    # TC 컬럼 (흐름트리 오른쪽에 1:1 배치)
    if include_tc and path_tracker:
        tc_start_col = state["max_col"] + 3  # 두 칸 여백 후 TC 시작
        tc_headers = ["TC ID", "사전조건", "테스트 스텝", "기대결과", "기획서 페이지"]
        tc_fields = ["tc_id", "preconditions", "steps", "expected_result", "spec_page"]
        tc_widths = [10, 30, 50, 30, 10]

        # TC 헤더
        ws.cell(1, tc_start_col, "TC 목록").font = Font(bold=True, size=13)
        for j, (h, w) in enumerate(zip(tc_headers, tc_widths)):
            c = ws.cell(3, tc_start_col + j, h)
            c.font = Font(bold=True, size=9, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0f172a")
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(tc_start_col + j)].width = w

        # 구분 열 (흐름트리 ↔ TC 사이)
        ws.column_dimensions[get_column_letter(state["max_col"] + 2)].width = 2

        # TC 데이터: Gemini 생성 tc_data가 있으면 그것 사용, 없으면 기계적 선형화
        tc_counter = 1
        for row_num in sorted(path_tracker.keys()):
            if tc_data and row_num in tc_data:
                tc = tc_data[row_num]
                tc["tc_id"] = f"TC-{tc_counter:03d}"
            else:
                path = path_tracker[row_num]
                tc = _build_tc_from_path(path, f"TC-{tc_counter:03d}")
            tc_counter += 1
            for j, field in enumerate(tc_fields):
                val = tc.get(field, "")
                cell = ws.cell(row_num, tc_start_col + j, val)
                cell.font = Font(size=9)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

    # path_tracker 외부 반환 (Gemini TC 생성을 위해 export 엔드포인트에서 활용)
    if _out_paths is not None:
        _out_paths.update(path_tracker)

    ws.freeze_panes = "B4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def flow_tree_stats(flow_tree: Dict) -> Dict:
    """노드 타입 분포·최대 깊이·총 노드 수 (요약·검증용)."""
    counts: Dict[str, int] = {}
    max_depth = [0]
    total = [0]

    def walk(node: Dict, depth: int):
        total[0] += 1
        counts[node.get("type", "?")] = counts.get(node.get("type", "?"), 0) + 1
        max_depth[0] = max(max_depth[0], depth)
        for ch in node.get("children") or []:
            walk(ch, depth + 1)

    for root in (flow_tree or {}).get("tree", []) or []:
        walk(root, 0)
    return {"types": counts, "max_depth": max_depth[0], "total": total[0]}
