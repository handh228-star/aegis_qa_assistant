import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List
from pathlib import Path
from datetime import datetime
from app.models.testcase import TestCase, TCType, TCPriority


TYPE_LABEL = {
    TCType.POSITIVE: "정상",
    TCType.NEGATIVE: "비정상",
    TCType.BOUNDARY: "경계값",
    TCType.EXCEPTION: "예외",
}

PRIORITY_LABEL = {
    TCPriority.HIGH: "높음",
    TCPriority.MEDIUM: "중간",
    TCPriority.LOW: "낮음",
}

PRIORITY_COLOR = {
    TCPriority.HIGH: "FF4444",
    TCPriority.MEDIUM: "FFA500",
    TCPriority.LOW: "44AA44",
}


def generate_tc_report(testcases: List[TestCase], project_name: str, output_dir: Path) -> str:
    wb = openpyxl.Workbook()

    _create_summary_sheet(wb.active, testcases, project_name)
    _create_tc_sheet(wb.create_sheet("TC 목록"), testcases)

    filename = f"TC_Report_{project_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = output_dir / filename
    wb.save(output_path)
    return str(output_path)


def _header_style(cell, bg_color="2C3E50"):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(fill_type="solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _apply_border(cell)


def _apply_border(cell):
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _create_summary_sheet(ws, testcases: List[TestCase], project_name: str):
    ws.title = "요약"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20

    ws["A1"] = f"프로젝트: {project_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"] = f"총 TC 수: {len(testcases)}"
    ws["A3"].font = Font(bold=True)

    ws["A5"] = "유형별 분포"
    _header_style(ws["A5"])
    ws["B5"] = "건수"
    _header_style(ws["B5"])

    type_counts = {t: 0 for t in TCType}
    for tc in testcases:
        type_counts[tc.tc_type] += 1

    for row, (tc_type, count) in enumerate(type_counts.items(), start=6):
        ws.cell(row=row, column=1, value=TYPE_LABEL[tc_type])
        ws.cell(row=row, column=2, value=count)
        _apply_border(ws.cell(row=row, column=1))
        _apply_border(ws.cell(row=row, column=2))

    ws["A11"] = "우선순위별 분포"
    _header_style(ws["A11"])
    ws["B11"] = "건수"
    _header_style(ws["B11"])

    priority_counts = {p: 0 for p in TCPriority}
    for tc in testcases:
        priority_counts[tc.priority] += 1

    for row, (priority, count) in enumerate(priority_counts.items(), start=12):
        ws.cell(row=row, column=1, value=PRIORITY_LABEL[priority])
        ws.cell(row=row, column=2, value=count)
        _apply_border(ws.cell(row=row, column=1))
        _apply_border(ws.cell(row=row, column=2))


def _create_tc_sheet(ws, testcases: List[TestCase]):
    headers = ["TC ID", "카테고리", "제목", "유형", "우선순위", "사전조건", "테스트 단계", "기대 결과"]
    col_widths = [10, 18, 35, 10, 10, 25, 45, 35]

    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        _header_style(cell)
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 25

    for row, tc in enumerate(testcases, start=2):
        preconditions = "\n".join(tc.preconditions) if tc.preconditions else ""
        steps_text = "\n".join(
            [f"{s['step']}. {s['action']}" for s in tc.steps] if tc.steps else []
        )

        values = [
            tc.tc_id,
            tc.category,
            tc.title,
            TYPE_LABEL.get(tc.tc_type, tc.tc_type),
            PRIORITY_LABEL.get(tc.priority, tc.priority),
            preconditions,
            steps_text,
            tc.expected_result,
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            _apply_border(cell)

        priority_cell = ws.cell(row=row, column=5)
        color = PRIORITY_COLOR.get(tc.priority, "000000")
        priority_cell.font = Font(bold=True, color=color)

        ws.row_dimensions[row].height = max(40, len(steps_text.split("\n")) * 15)
