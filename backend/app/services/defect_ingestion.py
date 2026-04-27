import openpyxl
from pathlib import Path
from typing import List, Dict
from app.core.config import settings
from app.services.manual_ingestion import _get_embedding_fn, SimpleVectorStore


def get_defect_collection() -> SimpleVectorStore:
    return SimpleVectorStore(settings.VECTOR_DB_DIR, _get_embedding_fn(), name="defect_history")


def _find_col(headers: List[str], *keywords) -> int:
    for i, h in enumerate(headers):
        if h and any(k in h for k in keywords):
            return i
    return -1


def _extract_defects_from_sheet(sheet, source_file: str) -> List[Dict]:
    defects = []
    headers = None
    col_result = col_mid = col_sub = col_step = col_note = col_jira = -1

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if not any(row):
            continue

        # 헤더 행 탐지
        if headers is None:
            row_strs = [str(c).strip() if c is not None else "" for c in row]
            if any("검증결과" in s or "결과" in s for s in row_strs):
                headers = row_strs
                col_result = _find_col(headers, "검증결과", "결과")
                col_mid    = _find_col(headers, "중분류")
                col_sub    = _find_col(headers, "소분류")
                col_step   = _find_col(headers, "시험절차", "절차")
                col_note   = _find_col(headers, "비고")
                col_jira   = _find_col(headers, "Jira", "jira", "이슈")
            continue

        if col_result == -1:
            continue

        result_val = str(row[col_result]).strip() if row[col_result] is not None else ""
        if result_val.lower() != "fail":
            continue

        def get(col):
            return str(row[col]).strip() if col >= 0 and col < len(row) and row[col] is not None else ""

        note = get(col_note)
        step = get(col_step)
        if not note and not step:
            continue

        defects.append({
            "source": source_file,
            "mid_category": get(col_mid),
            "category": get(col_sub),
            "step": step,
            "note": note,
            "jira": get(col_jira),
            "row_idx": row_idx,
        })

    return defects


def _format_defect_doc(d: Dict) -> str:
    parts = []
    area = " > ".join(filter(None, [d["mid_category"], d["category"]]))
    if area:
        parts.append(f"기능영역: {area}")
    if d["note"]:
        parts.append(f"결함내용: {d['note']}")
    if d["step"]:
        parts.append(f"재현절차: {d['step']}")
    if d["jira"]:
        parts.append(f"이슈: {d['jira']}")
    return "\n".join(parts)


def ingest_test_result(excel_path: str) -> Dict:
    """테스트 결과 Excel에서 Fail TC를 결함 이력 DB에 저장"""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    source_file = Path(excel_path).stem
    collection = get_defect_collection()

    existing = collection.get(where={"source": source_file})
    if existing["ids"]:
        collection.delete(ids=existing["ids"])
        print(f"[결함이력] 기존 {len(existing['ids'])}건 삭제")

    all_defects = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        defects = _extract_defects_from_sheet(sheet, source_file)
        if defects:
            print(f"  시트 '{sheet_name}': {len(defects)}건 결함 추출")
        all_defects.extend(defects)

    if not all_defects:
        return {"source": source_file, "defects": 0, "status": "no_defects"}

    batch_size = 50
    for i in range(0, len(all_defects), batch_size):
        batch = all_defects[i:i + batch_size]
        collection.add(
            ids=[f"{source_file}_defect_{d['row_idx']}" for d in batch],
            documents=[_format_defect_doc(d) for d in batch],
            metadatas=[{
                "source": d["source"],
                "category": d["category"],
                "mid_category": d["mid_category"],
                "jira": d["jira"],
            } for d in batch],
        )

    print(f"[결함이력] '{source_file}' 총 {len(all_defects)}건 저장 완료")
    return {"source": source_file, "defects": len(all_defects), "status": "success"}


def get_defect_stats() -> Dict:
    collection = get_defect_collection()
    all_data = collection.get()

    stats: Dict[str, int] = {}
    for meta in all_data["metadatas"]:
        src = meta.get("source", "unknown")
        stats[src] = stats.get(src, 0) + 1

    return {
        "total_defects": collection.count(),
        "sources": [{"source": k, "defects": v} for k, v in sorted(stats.items())],
    }
