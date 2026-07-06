import shutil
import uuid
import json
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import List, Optional, Any
from datetime import datetime
from pathlib import Path
from app.models.database import get_db
from app.models.document import Document, DocumentStatus
from app.models.testcase import TestCase, TCType, TCPriority, TCStatus, ChangeType
from app.models.project import Project
from app.models.qa_ruleset import QARuleSet
from app.core.config import settings
from app.services.document_parser import get_pdf_page_count
from app.services.tc_generator import generate_tc_from_pdf, generate_tc_from_tree, build_menu_tree, build_flow_tree, linearize_flow_tree, check_flow_coverage, repair_flow_tree, generate_tcs_from_flow_paths
from app.services.flow_tree_report import render_flow_tree_excel, flow_tree_stats

router = APIRouter(prefix="/documents", tags=["documents"])


class DocumentResponse(BaseModel):
    id: int
    project_id: int
    original_filename: str
    total_pages: int
    tc_level: int = 3
    status: str
    progress_current: int = 0
    progress_total: int = 0
    tc_started_at: Optional[datetime] = None
    created_at: datetime

    class Config:
        from_attributes = True


_VALID_TC_TYPES = {"positive", "negative", "boundary", "exception"}
_VALID_PRIORITIES = {"high", "medium", "low"}
_VALID_CHANGE_TYPES = {"new_feature", "modification", "bug_fix", "unknown"}


def _save_testcases(db, doc, result):
    for i, tc_data in enumerate(result["testcases"], start=1):
        tc_type_val = tc_data.get("tc_type", "positive")
        if tc_type_val not in _VALID_TC_TYPES:
            tc_type_val = "positive"

        priority_val = tc_data.get("priority", "medium")
        if priority_val not in _VALID_PRIORITIES:
            priority_val = "medium"

        change_type_val = tc_data.get("change_type", "unknown")
        if change_type_val not in _VALID_CHANGE_TYPES:
            change_type_val = "unknown"

        tc = TestCase(
            document_id=doc.id,
            tc_id=tc_data.get("tc_id", f"TC-{i:03d}"),
            category=tc_data.get("category", "기타"),
            title=tc_data.get("title", ""),
            objective=tc_data.get("objective", ""),
            spec_page=(str(tc_data.get("spec_page") or "").strip() or None),
            preconditions=tc_data.get("preconditions", []),
            steps=tc_data.get("steps", []),
            expected_result=tc_data.get("expected_result", ""),
            tc_type=TCType(tc_type_val),
            priority=TCPriority(priority_val),
            change_type=change_type_val,
            status=TCStatus.DRAFT,
        )
        db.add(tc)


def _get_ruleset(db, document_id: int):
    """문서의 프로젝트에 연결된 룰셋 반환 (없으면 기본 룰셋)"""
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        return None
    project = db.query(Project).filter(Project.id == doc.project_id).first()
    if project and project.ruleset_id:
        return db.query(QARuleSet).filter(QARuleSet.id == project.ruleset_id).first()
    return db.query(QARuleSet).filter(QARuleSet.is_default == True).first()


def _get_composed_ruleset(db, document_id: int):
    """시스템 룰셋(기본값) + 프로젝트 룰셋(추가 레이어)을 합성한 룰셋 객체를 반환.

    flow_rules는 시스템 룰셋이 항상 기반(Base)이 된다.
    - 내가(Claude) 시스템 룰셋을 업데이트하면 모든 프로젝트에 즉시 반영된다.
    - 프로젝트 룰셋에 추가된 내용이 있으면 기반 위에 레이어로 덧붙인다.
    tree_rules/tc_rules는 프로젝트 룰셋 우선, 없으면 시스템 룰셋.
    """
    from app.models.qa_ruleset import DEFAULT_FLOW_RULES
    from types import SimpleNamespace

    system_rs = db.query(QARuleSet).filter(QARuleSet.is_system == True).first()
    project_rs = _get_ruleset(db, document_id)

    system_flow = (system_rs.flow_rules if system_rs else None) or DEFAULT_FLOW_RULES

    # 프로젝트 룰셋이 시스템 룰셋과 다를 때만 추가 레이어로 붙인다 (단순 클론이면 생략)
    project_flow_extra = ""
    if project_rs and not project_rs.is_system and project_rs.flow_rules:
        if project_rs.flow_rules.strip() != system_flow.strip():
            project_flow_extra = project_rs.flow_rules

    composed_flow = (
        system_flow + "\n\n[프로젝트 추가 규칙]\n" + project_flow_extra
        if project_flow_extra else system_flow
    )

    return SimpleNamespace(
        name=(system_rs.name if system_rs else "시스템 룰셋"),
        flow_rules=composed_flow,
        tree_rules=(project_rs.tree_rules if project_rs else None) or (system_rs.tree_rules if system_rs else None),
        tc_rules=(project_rs.tc_rules if project_rs else None) or (system_rs.tc_rules if system_rs else None),
    )


def _analyze_document(document_id: int):
    """백그라운드에서 메뉴트리 생성"""
    from app.models.database import SessionLocal
    db = SessionLocal()
    try:
        doc = db.query(Document).filter(Document.id == document_id).first()
        if not doc:
            return
        doc.status = DocumentStatus.ANALYZING
        db.commit()

        ruleset = _get_ruleset(db, document_id)
        tree = build_menu_tree(doc.file_path, ruleset=ruleset, original_filename=doc.original_filename)
        doc.menu_tree = json.dumps(tree, ensure_ascii=False)

        # 상태 매트릭스(권한·코드 상태·이력 등 도메인 상태 차원) 추출 — 실패해도 진행
        from app.services.tc_generator import extract_state_inventory
        try:
            state_inv = extract_state_inventory(
                doc.file_path,
                original_filename=doc.original_filename,
                ruleset=ruleset,
            )
            doc.state_inventory = json.dumps(state_inv, ensure_ascii=False)
            n_dims = len(state_inv.get("state_dimensions", []))
            print(f"[상태 매트릭스] {n_dims}개 차원 도출")
        except Exception as e:
            print(f"[상태 매트릭스 추출 실패 — 무시] {e}")
            doc.state_inventory = json.dumps({"state_dimensions": []}, ensure_ascii=False)

        doc.status = DocumentStatus.ANALYZED
        doc.error_message = None
        db.commit()
        print(f"[메뉴트리 완료] doc_id={document_id}, 룰셋: {ruleset.name if ruleset else '없음'}")
    except Exception as e:
        print(f"[메뉴트리 실패] {e}")
        doc = db.query(Document).filter(Document.id == document_id).first()
        if doc:
            doc.status = DocumentStatus.FAILED
            doc.error_message = str(e)
            db.commit()
    finally:
        db.close()


def _process_document(document_id: int, tree_json: str = None):
    """백그라운드에서 TC 생성 처리 (기능 단위 재시도 포함)"""
    from app.models.database import SessionLocal
    db = SessionLocal()
    try:
        doc = db.query(Document).filter(Document.id == document_id).first()
        if not doc:
            return

        tc_level = doc.tc_level or 3
        doc.status = DocumentStatus.TC_GENERATING
        doc.progress_current = 0
        doc.progress_total = 0
        doc.tc_started_at = datetime.utcnow()
        db.commit()

        def update_progress(current, total, tc_count):
            try:
                _db = SessionLocal()
                _doc = _db.query(Document).filter(Document.id == document_id).first()
                if _doc:
                    _doc.progress_current = current
                    _doc.progress_total = total
                    _db.commit()
                _db.close()
            except Exception:
                pass

        ruleset = _get_ruleset(db, document_id)
        state_inventory = None
        if doc.state_inventory:
            try:
                state_inventory = json.loads(doc.state_inventory)
            except Exception:
                state_inventory = None

        if doc.menu_tree:
            tree = json.loads(doc.menu_tree)
            result = generate_tc_from_tree(
                tree, tc_level=tc_level, ruleset=ruleset,
                on_progress=update_progress, state_inventory=state_inventory,
            )
        else:
            result = generate_tc_from_pdf(doc.file_path, tc_level=tc_level)

        if not result.get("testcases"):
            raise ValueError("생성된 TC가 없습니다. 문서 분석 결과를 확인하세요.")

        _save_testcases(db, doc, result)
        doc.status = DocumentStatus.TC_GENERATED
        doc.error_message = None
        db.commit()

        try:
            from app.services.tc_ingestion import ingest_testcases
            ingest_testcases(doc.id, result["testcases"])
        except Exception as e:
            print(f"[TC 이력 저장 실패] {e}")

    except Exception as e:
        print(f"[TC생성 실패] {e}")
        doc = db.query(Document).filter(Document.id == document_id).first()
        if doc:
            doc.status = DocumentStatus.FAILED
            doc.error_message = str(e)
            db.commit()
    finally:
        db.close()


@router.post("/{project_id}/upload", response_model=DocumentResponse)
async def upload_document(
    project_id: int,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    tc_level: int = Form(3),
    db: Session = Depends(get_db),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다")

    if not file.filename.endswith(".pdf"):
        raise HTTPException(status_code=400, detail="PDF 파일만 업로드 가능합니다")

    unique_name = f"{uuid.uuid4().hex}_{file.filename}"
    file_path = settings.UPLOAD_DIR / unique_name

    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    page_count = get_pdf_page_count(str(file_path))

    doc = Document(
        project_id=project_id,
        filename=unique_name,
        original_filename=file.filename,
        file_path=str(file_path),
        total_pages=page_count,
        tc_level=max(1, min(5, tc_level)),
        status=DocumentStatus.UPLOADED,
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)

    # 메뉴트리 먼저 생성 (사용자 검토 후 TC 생성)
    background_tasks.add_task(_analyze_document, doc.id)

    return doc


@router.get("/{document_id}/tree")
def get_tree(document_id: int, db: Session = Depends(get_db)):
    """메뉴트리 조회"""
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다")
    if not doc.menu_tree:
        raise HTTPException(status_code=404, detail="메뉴트리가 아직 생성되지 않았습니다")
    return json.loads(doc.menu_tree)


@router.put("/{document_id}/tree")
def update_tree(document_id: int, tree: Any, db: Session = Depends(get_db)):
    """사용자가 수정한 메뉴트리 저장"""
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다")
    doc.menu_tree = json.dumps(tree, ensure_ascii=False)
    db.commit()
    return {"message": "저장되었습니다"}


@router.get("/{document_id}/state-inventory")
def get_state_inventory(document_id: int, db: Session = Depends(get_db)):
    """상태 매트릭스(권한·코드 상태·이력 등 도메인 상태 차원) 조회"""
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다")
    if not doc.state_inventory:
        return {"state_dimensions": []}
    try:
        return json.loads(doc.state_inventory)
    except Exception:
        return {"state_dimensions": []}


@router.get("/{document_id}/tree/export")
def export_tree_excel(document_id: int, db: Session = Depends(get_db)):
    """메뉴트리를 Excel로 내보내기"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc or not doc.menu_tree:
        raise HTTPException(status_code=404, detail="메뉴트리가 없습니다")

    tree_data = json.loads(doc.menu_tree)

    # 리프 노드 ↔ TC 매핑 (category = 리프 전체 경로). 삭제 예정 TC 제외.
    tcs = (
        db.query(TestCase)
        .filter(TestCase.document_id == document_id)
        .filter((TestCase.review_status != "deleted") | (TestCase.review_status.is_(None)))
        .order_by(TestCase.id)
        .all()
    )
    tc_by_path = {}
    for tc in tcs:
        tc_by_path.setdefault(tc.category, []).append(tc)

    CHANGE_LABEL = {
        "new_feature": "신규", "modification": "수정",
        "bug_fix": "버그수정", "unknown": "일반",
    }
    TC_TYPE_LABEL = {"positive": "정상", "negative": "비정상",
                     "boundary": "경계값", "exception": "예외"}
    PRIORITY_LABEL = {"high": "높음", "medium": "중간", "low": "낮음"}
    PRIORITY_COLOR = {"high": "dc2626", "medium": "d97706", "low": "16a34a"}
    TC_ROW_FILL = "f9fafb"

    def _enum_val(v):
        return getattr(v, "value", v)
    DEPTH_STYLES = [
        {"fill": "1e40af", "font_color": "FFFFFF", "font_size": 12, "bold": True},   # 1단계
        {"fill": "3b82f6", "font_color": "FFFFFF", "font_size": 11, "bold": True},   # 2단계
        {"fill": "bfdbfe", "font_color": "1e3a8a", "font_size": 11, "bold": False},  # 3단계
        {"fill": "eff6ff", "font_color": "1e3a8a", "font_size": 10, "bold": False},  # 4단계
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "메뉴트리"

    thin = Side(style="thin", color="d1d5db")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["계층", "기능명", "변경유형", "기획서페이지", "설명", "검증포인트"]
    col_widths = [8, 40, 10, 12, 40, 50]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="374151")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    row_num = 2

    def write_tc_row(tc, depth):
        """리프 노드 아래에 매핑된 TC를 하위 행으로 펼친다."""
        nonlocal row_num
        indent = "    " * depth
        prio = _enum_val(tc.priority)
        values = [
            "TC",
            f"{indent}└ [{tc.tc_id}] {tc.title}",
            TC_TYPE_LABEL.get(_enum_val(tc.tc_type), ""),
            str(tc.spec_page or "").strip(),
            tc.objective or "",
            tc.expected_result or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=TC_ROW_FILL)
            color = PRIORITY_COLOR.get(prio, "374151") if col == 2 else "6b7280"
            cell.font = Font(size=9, color=color, bold=(col == 2 and prio == "high"))
            cell.alignment = Alignment(vertical="top", wrap_text=True,
                                       horizontal="center" if col in (1, 3, 4) else "left")
            cell.border = border
        longest = max(len(tc.objective or ""), len(tc.expected_result or ""), 1)
        ws.row_dimensions[row_num].height = max(16, (longest // 28 + 1) * 14)
        row_num += 1

    def write_node(node, depth=0, parent_path=""):
        nonlocal row_num
        style = DEPTH_STYLES[min(depth, len(DEPTH_STYLES) - 1)]
        indent = "  " * depth
        name = node.get("name", "")
        current_path = f"{parent_path} > {name}" if parent_path else name
        key_points = "\n".join(f"• {kp}" for kp in (node.get("key_points") or []))
        children = node.get("children") or []

        values = [
            depth + 1,
            indent + name,
            CHANGE_LABEL.get(node.get("change_type", "unknown"), "일반"),
            str(node.get("spec_page") or "").strip(),
            node.get("description", ""),
            key_points,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=style["fill"])
            cell.font = Font(bold=style["bold"], color=style["font_color"], size=style["font_size"])
            cell.alignment = Alignment(vertical="center", wrap_text=True,
                                       horizontal="center" if col in (1, 4) else "left")
            cell.border = border

        if key_points:
            ws.row_dimensions[row_num].height = max(18, len(key_points.splitlines()) * 16)
        else:
            ws.row_dimensions[row_num].height = 20

        row_num += 1

        if children:
            for child in children:
                write_node(child, depth + 1, current_path)
        else:
            # 리프 노드: 이 노드 경로에 매핑된 실제 TC를 하위 행으로 펼침
            for tc in tc_by_path.get(current_path, []):
                write_tc_row(tc, depth + 1)

    for node in tree_data.get("tree", []):
        write_node(node)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"menutree_{document_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ============================================================================
# 흐름 트리(Flow Tree) — 행동 흐름 메뉴트리. menu_tree(구조적)와 별도 보관.
# ============================================================================
def _build_flow_tree_bg(document_id: int):
    """백그라운드: PDF에서 흐름 트리 추출 → doc.flow_tree에 저장.

    구조적 트리(menu_tree)가 아직 없으면 먼저 추출해 화면 골격으로 삼는다 — 같은 화면이
    흐름 트리에서 별개 루트로 중복 생성되는 것을 막기 위함(설계: 구조적 트리 → 흐름 트리 2단계).
    """
    from app.models.database import SessionLocal
    db = SessionLocal()
    try:
        doc = db.query(Document).filter(Document.id == document_id).first()
        if not doc:
            return
        # 시스템 룰셋(기반) + 프로젝트 룰셋(추가 레이어) 합성 — flow_rules는 항상 시스템이 기반
        ruleset = _get_composed_ruleset(db, document_id)
        # menu_tree 추출은 기존 방식(프로젝트 룰셋)으로 그대로
        raw_ruleset = _get_ruleset(db, document_id)

        if not doc.menu_tree:
            print(f"[흐름트리] 구조적 트리 없음 — 먼저 추출 doc_id={document_id}")
            menu_tree = build_menu_tree(doc.file_path, ruleset=raw_ruleset, original_filename=doc.original_filename)
            doc.menu_tree = json.dumps(menu_tree, ensure_ascii=False)
            db.commit()
        else:
            menu_tree = json.loads(doc.menu_tree)

        flow = build_flow_tree(
            doc.file_path, ruleset=ruleset, original_filename=doc.original_filename, menu_tree=menu_tree,
        )
        st = flow_tree_stats(flow)
        print(f"[흐름트리 1차] doc_id={document_id}, 노드 {st['total']}개, 깊이 {st['max_depth']}")

        # 자동 교정 패스: 1차 결과를 커버리지 점검 후 not_followed 위반을 Gemini로 수정
        from app.models.qa_ruleset import DEFAULT_FLOW_RULES
        flow_rules = (ruleset.flow_rules if ruleset and ruleset.flow_rules else DEFAULT_FLOW_RULES)
        tree_rules = (ruleset.tree_rules if ruleset and ruleset.tree_rules else "")
        sections = []
        if flow_rules:
            sections.append(f"[target=flow]\n{flow_rules}")
        if tree_rules:
            sections.append(f"[target=tree]\n{tree_rules}")
        rules_text = "\n\n".join(sections)
        check_result = check_flow_coverage(flow, rules_text)
        findings = check_result.get("findings", [])
        fixable_count = sum(1 for f in findings if f.get("fixability") == "not_followed")
        if fixable_count > 0:
            print(f"[흐름트리 교정 시작] not_followed {fixable_count}건 기획서 참조 교정 중...")
            flow = repair_flow_tree(flow, findings,
                                    pdf_path=doc.file_path,
                                    original_filename=doc.original_filename)
            st = flow_tree_stats(flow)
            print(f"[흐름트리 교정 완료] 노드 {st['total']}개")
        else:
            print(f"[흐름트리 교정 불필요] not_followed 위반 없음 (spec_limited {len(findings) - fixable_count}건은 건너뜀)")

        doc.flow_tree = json.dumps(flow, ensure_ascii=False)
        db.commit()
        print(f"[흐름트리 완료] doc_id={document_id}, 노드 {st['total']}개, 깊이 {st['max_depth']}, 타입 {st['types']}")
    except Exception as e:
        err_msg = str(e)[:300]
        print(f"[흐름트리 실패] doc_id={document_id}: {err_msg}")
        try:
            # 실패 상태를 DB에 기록해서 프론트엔드 폴링이 인식하게 함
            from app.models.database import SessionLocal as _SL
            _db = _SL()
            _doc = _db.query(Document).filter(Document.id == document_id).first()
            if _doc:
                _doc.flow_tree = '{"error": true, "message": "' + err_msg.replace('"', "'") + '"}'
                _db.commit()
            _db.close()
        except Exception:
            pass
    finally:
        db.close()


@router.post("/{document_id}/flow-tree")
def start_build_flow_tree(document_id: int, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    """흐름 트리 추출 시작 (백그라운드). 완료 후 GET /flow-tree 로 조회."""
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다")
    # 기존 트리를 즉시 지워야 폴링이 "추출 중(ready=False)" 상태를 올바르게 인식한다.
    # 지우지 않으면 첫 번째 폴링에서 ready=True가 바로 반환돼 직전 트리를 새 결과로 오인한다.
    doc.flow_tree = None
    db.commit()
    background_tasks.add_task(_build_flow_tree_bg, document_id)
    return {"message": "흐름 트리 추출을 시작했습니다", "document_id": document_id}


@router.get("/{document_id}/flow-tree")
def get_flow_tree(document_id: int, db: Session = Depends(get_db)):
    """저장된 흐름 트리 JSON 조회 (아직 없으면 ready=False)."""
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다")
    if not doc.flow_tree:
        return {"ready": False}
    flow = json.loads(doc.flow_tree)
    if flow.get("error"):
        return {"ready": False, "error": flow.get("message", "알 수 없는 오류")}
    return {"ready": True, "flow_tree": flow, "stats": flow_tree_stats(flow)}


@router.get("/{document_id}/flow-tree/export")
def export_flow_tree_excel(document_id: int, with_tc: bool = False, db: Session = Depends(get_db)):
    """흐름 트리를 QA 양식 Excel 그리드로 내보내기.

    with_tc=true이면:
      1단계) 흐름트리 렌더링 + 경로(path_tracker) 수집
      2단계) Gemini로 자연어 TC 생성 (기획서 참조, 1~2분 추가)
      3단계) TC를 흐름트리 오른쪽에 1:1 배치해 다운로드
    """
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc or not doc.flow_tree:
        raise HTTPException(status_code=404, detail="흐름 트리가 없습니다")

    if with_tc:
        # 1단계: 흐름트리 렌더링 + 경로 수집
        out_paths: dict = {}
        buf_plain = render_flow_tree_excel(json.loads(doc.flow_tree), _out_paths=out_paths)

        # 2단계: Gemini 자연어 TC 생성 (PDF 없이 노드 content 기반, 배치 처리)
        tc_by_row = generate_tcs_from_flow_paths(out_paths)

        # 3단계: TC 포함 최종 Excel 생성
        buf = render_flow_tree_excel(
            json.loads(doc.flow_tree),
            include_tc=True,
            tc_data=tc_by_row,
        )
        fname = f"flowtree_tc_{document_id}.xlsx"
    else:
        buf = render_flow_tree_excel(json.loads(doc.flow_tree))
        fname = f"flowtree_{document_id}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.post("/{document_id}/flow-tree/coverage-check")
def flow_coverage_check(document_id: int, db: Session = Depends(get_db)):
    """흐름 트리를 프로젝트 룰셋(tree_rules + flow_rules)과 대조해 누락·위반을 점검한다.

    룰셋을 능동적으로 활용하는 게이트. 결과(findings)는 재추출 시 룰셋에 반영할 단서가 된다.
    """
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc or not doc.flow_tree:
        raise HTTPException(status_code=404, detail="흐름 트리가 없습니다. 먼저 흐름 트리를 추출하세요.")
    # 생성 때와 동일하게 합성 룰셋 사용 — 시스템 기반 + 프로젝트 추가 레이어
    ruleset = _get_composed_ruleset(db, document_id)
    flow_rules = ruleset.flow_rules or ""
    tree_rules = ruleset.tree_rules or ""
    # 섹션 라벨을 붙여서 LLM이 각 finding의 출처(target)를 구분할 수 있게 한다 — "+ 룰셋에 강화
    # 규칙 추가" 버튼이 올바른 필드(flow_rules/tree_rules)에 저장하려면 이 구분이 필요하다.
    sections = []
    if flow_rules:
        sections.append(f"[target=flow — 흐름 트리 구조 문법 규칙]\n{flow_rules}")
    if tree_rules:
        sections.append(f"[target=tree — 관점 가이드(커버리지)]\n{tree_rules}")
    rules_text = "\n\n".join(sections)
    result = check_flow_coverage(json.loads(doc.flow_tree), rules_text)
    return {
        "ruleset": getattr(ruleset, "name", None),
        "findings": result.get("findings", []),
        "note": result.get("note"),
        "error": result.get("error"),
    }


class RuleAppendIn(BaseModel):
    rule: str
    target: Optional[str] = "tree"   # 'tree' | 'tc' | 'flow'
    scope: Optional[str] = "project"  # 'project'(이 프로젝트에만) | 'system'(시스템 룰셋에 반영 — 전체 프로젝트 적용)


@router.post("/{document_id}/ruleset/append-rule")
def append_ruleset_rule(document_id: int, payload: RuleAppendIn, db: Session = Depends(get_db)):
    """룰셋에 규칙을 직접 추가한다 (프로그래매틱 용도).

    scope='system': 시스템 룰셋에 추가. scope 미지정: 이 프로젝트 룰셋에 추가.
    클론-온-라이트 없음 — 프로젝트는 명시적으로 선택된 룰셋만 사용한다.
    """
    rule = (payload.rule or "").strip()
    if not rule:
        raise HTTPException(status_code=400, detail="규칙 내용이 비어 있습니다")
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다")

    field = {"tc": "tc_rules", "flow": "flow_rules"}.get(payload.target, "tree_rules")

    import re as _re
    def _norm(s):
        return _re.sub(r"\s+", "", (s or "")).lower()

    if payload.scope == "system":
        rs = db.query(QARuleSet).filter(QARuleSet.is_system == True).first()
        if not rs:
            raise HTTPException(status_code=404, detail="시스템 룰셋을 찾을 수 없습니다")
    else:
        rs = _get_ruleset(db, document_id)
        if not rs:
            rs = db.query(QARuleSet).filter(QARuleSet.is_system == True).first()

    current_text = (getattr(rs, field) if rs else "") or ""
    if _norm(rule) and _norm(rule) in _norm(current_text):
        return {"ruleset_id": rs.id if rs else None, "ruleset_name": rs.name if rs else None,
                "duplicate": True, "message": "이미 룰셋에 반영된 내용입니다."}

    existing = (getattr(rs, field) or "").rstrip()
    stamp = datetime.now().strftime("%Y-%m-%d")
    setattr(rs, field, f"{existing}\n- (QA 반영 {stamp}) {rule}")
    db.commit()
    return {
        "ruleset_id": rs.id, "ruleset_name": rs.name, "duplicate": False,
        "target": payload.target or "tree", "scope": payload.scope or "project",
        "message": ("시스템 룰셋에 추가했습니다." if payload.scope == "system" else "룰셋에 추가했습니다."),
    }


@router.post("/{document_id}/flow-tree/generate-tc")
def generate_tc_from_flow(document_id: int, db: Session = Depends(get_db)):
    """흐름 트리를 선형화해 TC 생성 (트리=마스터, TC=파생).

    기존 TC를 모두 교체한다(흐름 트리에서 재파생). 흐름 트리가 없으면 404.
    """
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc or not doc.flow_tree:
        raise HTTPException(status_code=404, detail="흐름 트리가 없습니다. 먼저 흐름 트리를 추출하세요.")
    result = linearize_flow_tree(json.loads(doc.flow_tree))
    db.query(TestCase).filter(TestCase.document_id == document_id).delete()
    _save_testcases(db, doc, result)
    doc.status = DocumentStatus.TC_GENERATED
    db.commit()
    return {"message": "흐름 트리에서 TC를 생성했습니다", "count": len(result["testcases"])}


@router.post("/{document_id}/generate-tc")
def start_generate_tc(document_id: int, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    """메뉴트리 검토 완료 후 TC 생성 시작"""
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다")
    if doc.status == DocumentStatus.TC_GENERATING:
        raise HTTPException(status_code=409, detail="이미 TC 생성 중입니다")

    # 기존 TC 삭제 (재생성 시)
    db.query(TestCase).filter(TestCase.document_id == document_id).delete()
    db.commit()

    background_tasks.add_task(_process_document, document_id)
    return {"message": "TC 생성을 시작했습니다", "document_id": document_id}


@router.get("/{project_id}/", response_model=List[DocumentResponse])
def list_documents(project_id: int, db: Session = Depends(get_db)):
    return db.query(Document).filter(Document.project_id == project_id).all()


@router.get("/status/{document_id}")
def get_document_status(document_id: int, db: Session = Depends(get_db)):
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다")

    tc_count = db.query(TestCase).filter(TestCase.document_id == document_id).count()
    return {
        "id": doc.id,
        "status": doc.status,
        "total_pages": doc.total_pages,
        "tc_level": doc.tc_level or 3,
        "tc_count": tc_count,
        "error_message": doc.error_message,
        "progress_current": doc.progress_current or 0,
        "progress_total": doc.progress_total or 0,
        "tc_started_at": doc.tc_started_at.isoformat() if doc.tc_started_at else None,
    }
